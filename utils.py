import pandas as pd
from openpyxl import load_workbook


def read_input_file(filename):
    """Parse a .xlsx input file"""

    wb = load_workbook(filename=filename)
    sheet_names = wb.sheetnames

    for sn in ("costs", "timeseries", "settings"):
        if sn not in sheet_names:
            raise ValueError(
                f"The sheet '{sn}' is missing in your input file {filename}"
            )

    name = "costs"
    headers = [c.value for c in wb[name][1]]
    df = pd.DataFrame(tuple(wb[name].values)[1:], columns=headers)
    # drop the lines which do not define a new asset
    df = df.loc[df.asset.notna()]
    # drop the columns with no headers
    df_costs = df[df.columns.dropna()].set_index("asset").fillna(0)

    for col_name in ("asset", "epc", "opex_variable"):
        if col_name not in headers:
            raise ValueError(
                f"The column header '{col_name}' is missing in your input file {filename} under the '{name}' sheet"
            )

    name = "timeseries"
    headers = [c.value for c in wb[name][1]]

    for hd in ("CriticalDemand", "Demand", "SolarGen"):
        if hd not in headers:
            raise ValueError(
                f"The column header '{hd}' is missing in your input file {filename} under the '{name}' sheet"
            )

    df_timeseries = pd.DataFrame(tuple(wb[name].values)[1:], columns=headers)

    name = "settings"
    headers = [c.value for c in wb[name][1]]

    for hd in ("param", "value"):
        if hd not in headers:
            raise ValueError(
                f"The column header '{hd}' is missing in your input file {filename} under the '{name}' sheet"
            )
    df = pd.DataFrame(tuple(wb[name].values)[1:], columns=headers)
    df = df[df.columns.dropna()].set_index("param")
    df_settings = df["value"]

    return df_costs, df_timeseries, df_settings
